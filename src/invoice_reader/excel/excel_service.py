"""Safe monthly Excel workbook operations."""

import os
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook, load_workbook

from invoice_reader.application.models import InvoiceRecord
from invoice_reader.application.sdr_amount import parse_sdr_amount
from invoice_reader.excel.excel_errors import DuplicatePlmnError, ExcelHeaderError, ExcelLockedError


EXCEL_HEADERS = ("PLMN", "Invoice No.", "SDR amount", "TAP start", "TAP end", "审批时间")


class ExcelService:
    """Create, validate, and atomically update a monthly invoice workbook."""

    def create_monthly_workbook(self, excel_path: str) -> None:
        """Create a workbook with the invoice header row."""
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Invoices"
        worksheet.append(EXCEL_HEADERS)
        self._configure_new_worksheet(worksheet)
        self._save_atomically(workbook, Path(excel_path))

    def validate_workbook(self, excel_path: str) -> None:
        """Ensure the active sheet starts with the required six headers."""
        workbook = load_workbook(excel_path, read_only=True, data_only=False)
        try:
            self._validate_headers(workbook.active)
        finally:
            workbook.close()

    def write_record(
        self,
        excel_path: str,
        record: InvoiceRecord,
        approved_at: str,
        overwrite: bool = False,
    ) -> None:
        """Append one record, or overwrite an existing exact-PLMN row."""
        workbook = load_workbook(excel_path)
        try:
            worksheet = workbook.active
            self._validate_headers(worksheet)
            row_number = self._find_plmn_row(worksheet, record.plmn.value)
            values = self._record_values(record, approved_at)
            if row_number is None:
                worksheet.append(values)
                row_number = worksheet.max_row
            elif overwrite:
                self._replace_row(worksheet, row_number, values)
            else:
                raise DuplicatePlmnError(self._row_values(worksheet, row_number))
            self._format_record_cells(worksheet, row_number)
            self._save_atomically(workbook, Path(excel_path))
        finally:
            workbook.close()

    def find_record(self, excel_path: str, plmn: str) -> tuple[str, ...] | None:
        """Return the exact PLMN row so a retry does not append it twice."""
        workbook = load_workbook(excel_path, read_only=True, data_only=False)
        try:
            worksheet = workbook.active
            self._validate_headers(worksheet)
            row_number = self._find_plmn_row(worksheet, plmn)
            return None if row_number is None else self._row_values(worksheet, row_number)
        finally:
            workbook.close()

    def read_plmns(self, excel_path: str) -> set[str]:
        """Return all non-empty PLMN identifiers in the active worksheet."""
        workbook = load_workbook(excel_path, read_only=True, data_only=False)
        try:
            worksheet = workbook.active
            self._validate_headers(worksheet)
            return {
                str(worksheet.cell(row=row_number, column=1).value).strip()
                for row_number in range(2, worksheet.max_row + 1)
                if worksheet.cell(row=row_number, column=1).value not in (None, "")
            }
        finally:
            workbook.close()

    def _validate_headers(self, worksheet: object) -> None:
        headers = tuple(worksheet.cell(row=1, column=index).value for index in range(1, 7))
        if headers != EXCEL_HEADERS:
            raise ExcelHeaderError("Excel 表头不匹配，请选择本程序创建的月度 Excel 文件。")

    def _configure_new_worksheet(self, worksheet: object) -> None:
        worksheet.freeze_panes = "A2"
        for column, width in zip(("A", "B", "C", "D", "E", "F"), (14, 22, 18, 14, 14, 26)):
            worksheet.column_dimensions[column].width = width

    def _find_plmn_row(self, worksheet: object, plmn: str) -> int | None:
        for row_number in range(2, worksheet.max_row + 1):
            if worksheet.cell(row=row_number, column=1).value == plmn:
                return row_number
        return None

    def _record_values(self, record: InvoiceRecord, approved_at: str) -> tuple[object, ...]:
        return (
            record.plmn.value,
            record.invoice_no.value,
            parse_sdr_amount(record.sdr_amount.value),
            self._tap_number(record.tap_start.value),
            self._tap_number(record.tap_end.value),
            approved_at,
        )

    def _tap_number(self, value: str) -> int | str:
        stripped_value = value.strip()
        if not stripped_value:
            return ""
        if not stripped_value.isdecimal():
            raise ValueError(f"TAP 不是有效整数：{value}")
        return int(stripped_value)

    def _format_record_cells(self, worksheet: object, row_number: int) -> None:
        for column_number in (2, 4, 5):
            worksheet.cell(row=row_number, column=column_number).number_format = "General"
        worksheet.cell(row=row_number, column=3).number_format = "0.00"

    def _replace_row(self, worksheet: object, row_number: int, values: tuple[object, ...]) -> None:
        for column_number, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column_number).value = value

    def _row_values(self, worksheet: object, row_number: int) -> tuple[str, ...]:
        return tuple(str(worksheet.cell(row=row_number, column=index).value or "") for index in range(1, 7))

    def _save_atomically(self, workbook: Workbook, destination: Path) -> None:
        destination.parent.mkdir(parents=True, exist_ok=True)
        temporary_path = self._temporary_path(destination)
        try:
            workbook.save(temporary_path)
            os.replace(temporary_path, destination)
        except PermissionError as error:
            raise ExcelLockedError("Excel 文件被占用，请关闭后重试。") from error
        finally:
            if temporary_path.exists():
                temporary_path.unlink()

    def _temporary_path(self, destination: Path) -> Path:
        with NamedTemporaryFile(
            prefix=f".{destination.stem}-",
            suffix=".xlsx",
            dir=destination.parent,
            delete=False,
        ) as temporary_file:
            return Path(temporary_file.name)
