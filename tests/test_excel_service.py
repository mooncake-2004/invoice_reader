"""Tests for monthly workbook creation and PLMN row updates."""

import pytest
from openpyxl import load_workbook

from invoice_reader.application.models import ExtractedField, InvoiceRecord
from invoice_reader.excel.excel_errors import DuplicatePlmnError, ExcelHeaderError
from invoice_reader.excel.excel_service import EXCEL_HEADERS, ExcelService


def _record(plmn: str = "ABCDE") -> InvoiceRecord:
    return InvoiceRecord(
        file_path="C:/invoices/ABCDE.pdf",
        plmn=ExtractedField(value=plmn),
        invoice_no=ExtractedField(value="INV-1001"),
        sdr_amount=ExtractedField(value="321.45"),
        tap_start=ExtractedField(value="100"),
        tap_end=ExtractedField(value="200"),
    )


def _cell(excel_path, cell_reference: str):
    workbook = load_workbook(excel_path, data_only=True)
    try:
        cell = workbook.active[cell_reference]
        return cell.value, cell.number_format
    finally:
        workbook.close()


def _rows(excel_path) -> list[tuple[object, ...]]:
    workbook = load_workbook(excel_path, data_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def test_creates_monthly_workbook_with_required_headers(tmp_path) -> None:
    excel_path = tmp_path / "2026-08.xlsx"
    ExcelService().create_monthly_workbook(str(excel_path))

    assert _rows(excel_path) == [EXCEL_HEADERS]


def test_rejects_a_workbook_with_different_headers(tmp_path) -> None:
    excel_path = tmp_path / "incorrect.xlsx"
    ExcelService().create_monthly_workbook(str(excel_path))
    workbook = load_workbook(excel_path)
    workbook.active["A1"] = "Other"
    workbook.save(excel_path)
    workbook.close()

    with pytest.raises(ExcelHeaderError):
        ExcelService().validate_workbook(str(excel_path))


def test_appends_then_overwrites_an_exact_plmn_row(tmp_path) -> None:
    excel_path = tmp_path / "2026-08.xlsx"
    service = ExcelService()
    service.create_monthly_workbook(str(excel_path))
    service.write_record(str(excel_path), _record(), "2026-08-06T10:00:00+08:00")

    with pytest.raises(DuplicatePlmnError):
        service.write_record(str(excel_path), _record(), "2026-08-06T11:00:00+08:00")

    updated = _record()
    updated.invoice_no.value = "INV-2002"
    service.write_record(str(excel_path), updated, "2026-08-06T11:00:00+08:00", overwrite=True)

    assert _rows(excel_path) == [
        EXCEL_HEADERS,
        ("ABCDE", "INV-2002", 321.45, 100, 200, "2026-08-06T11:00:00+08:00"),
    ]


def test_finds_an_existing_exact_plmn_row(tmp_path) -> None:
    excel_path = tmp_path / "2026-08.xlsx"
    service = ExcelService()
    service.create_monthly_workbook(str(excel_path))
    service.write_record(str(excel_path), _record(), "2026-08-06T10:00:00+08:00")

    assert service.find_record(str(excel_path), "ABCDE") == (
        "ABCDE",
        "INV-1001",
        "321.45",
        "100",
        "200",
        "2026-08-06T10:00:00+08:00",
    )


def test_writes_decimal_comma_sdr_as_a_numeric_excel_cell(tmp_path) -> None:
    excel_path = tmp_path / "2026-08.xlsx"
    service = ExcelService()
    service.create_monthly_workbook(str(excel_path))
    record = _record()
    record.sdr_amount.value = "0,23"

    service.write_record(str(excel_path), record, "2026-08-06T10:00:00+08:00")

    value, number_format = _cell(excel_path, "C2")
    assert value == 0.23
    assert isinstance(value, float)
    assert number_format == "0.00"


def test_writes_invoice_number_and_tap_cells_with_general_format(tmp_path) -> None:
    excel_path = tmp_path / "2026-08.xlsx"
    service = ExcelService()
    service.create_monthly_workbook(str(excel_path))
    record = _record()
    record.tap_start.value = "06338"
    record.tap_end.value = "06397"

    service.write_record(str(excel_path), record, "2026-08-06T10:00:00+08:00")

    invoice_number, invoice_format = _cell(excel_path, "B2")
    tap_start, tap_start_format = _cell(excel_path, "D2")
    tap_end, tap_end_format = _cell(excel_path, "E2")
    assert (invoice_number, invoice_format) == ("INV-1001", "General")
    assert (tap_start, tap_start_format) == (6338, "General")
    assert (tap_end, tap_end_format) == (6397, "General")
